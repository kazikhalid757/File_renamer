import os
import openpyxl
def create_excel_file(base_folder, add_to_sheet, start_row=2):
    for root, dirs, files in os.walk(base_folder):
        for dir_name in dirs:
            add_to_sheet.cell(row=start_row, column=1, value=os.path.join(root, dir_name))
            start_row += 1

    # Save the Excel file
    save_to_path = os.path.join(base_folder, "subfolders.xlsx")
    add_to_sheet.parent.save(save_to_path)
    return save_to_path

def collecting_sub_folder_data():
    # Function to create an Excel file with subfolder names recursively

    # Input the parent folder path
    parent_folder = input("Enter the path of the parent folder: ")

    # Create the Excel file with subfolder names recursively
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Subfolders"

    # Add headers
    worksheet.cell(row=1, column=1, value="Subfolder Name")
    worksheet.cell(row=1, column=2, value="MAIN")
    worksheet.cell(row=1, column=3, value="asin_number")

    excel_file = create_excel_file(parent_folder, worksheet)

    print(f"Excel file has been saved in '{excel_file}' location.")
    print("Please fill in the ASIN and MAIN columns as needed.")
    print("Once you've filled in the Excel file, run the utility again to rename the files.")
    print('\n')


def file_rename():
    # Function to read the Excel file and return folder paths and ASIN numbers
    def read_excel_file(excel_file_path):
        folder_data = []

        if not excel_file_path.endswith('.xlsx'):
            print("Invalid Excel file extension. Please provide a .xlsx file.")
            return folder_data

        if not os.path.isfile(excel_file_path):
            print("Excel file does not exist. Please provide a valid file path.")
            return folder_data

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3:
                    folder_path, main, asin_number = row[:3]  # Extract folder path, main_file_name, and ASIN number
                    folder_data.append((folder_path, main, asin_number))
                else:
                    print("Skipping row with insufficient data:", row)

        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")

        return folder_data

    # Function to rename files within a folder based on the provided condition
    def rename_files(folder_path, main_file_name, asin_number):
        if not os.path.exists(folder_path):
            print(f"Folder not found: {folder_path}")
            return

        pt_number = 1  # Initialize the part number

        # Get all file names in the subfolder
        file_names = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

        # Sort the file names based on name
        file_names.sort()

        for filename in file_names:
            base, ext = os.path.splitext(filename)
            possible_pt_number = base.rsplit("_", 1)[-1]
            if possible_pt_number.isnumeric():
                pt_number = int(possible_pt_number)
                if main_file_name is None:
                    pt_number -= 1

            if main_file_name is None:
                if base.upper().endswith("MAIN_1"):
                    new_filename = f"{asin_number}.MAIN{ext}"
                elif base.endswith("_1"):
                    new_filename = f"{asin_number}.MAIN{ext}"
                else:
                    new_filename = f"{asin_number}.PT{pt_number:02d}{ext}"
                    pt_number += 1  # Increment part number for non-main_file_name files
            elif filename == main_file_name:
                new_filename = f"{asin_number}.MAIN{ext}"
            else:
                new_filename = f"{asin_number}.PT{pt_number:02d}{ext}"
                pt_number += 1  # Increment part number for non-main_file_name files

            src_file_path = os.path.join(folder_path, filename)
            dst_file_path = os.path.join(folder_path, new_filename)

            if not os.path.exists(dst_file_path):
                os.rename(src_file_path, dst_file_path)
                print(f"Renamed: {filename} -> {new_filename}")

    # Input the path of the Excel file
    excel_file = input("Enter the full path of the Excel file (including filename and extension): ")

    folder_data = read_excel_file(excel_file)

    if folder_data:
        for folder_path, main_file_name, asin_number in folder_data:
            print(f"Processing folder: {folder_path}, ASIN: {asin_number}")
            rename_files(folder_path, main_file_name, asin_number)


num = int(input("Press 1 for collecting sub-folder data or 2 for file rename:"))
if num == 1:
    collecting_sub_folder_data()
else:
    file_rename()
